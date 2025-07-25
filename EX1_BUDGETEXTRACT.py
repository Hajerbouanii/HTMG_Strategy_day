# Install required packages:
# pip install openai PyMuPDF openpyxl
import fitz  # PyMuPDF
from openai import OpenAI
from openpyxl import Workbook, load_workbook

# Set up OpenAI client

import json
import os

client = OpenAI(
    api_key="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"
)


def generate_invoice_summary(text):
    prompt = (
        "Extract the invoice date, invoice company, invoice amount, currency, and summarize any issues or comments "
        "from the following text. Respond strictly in JSON format without markdown:\n"
        '{"invoice_date": "DD/MM/YYYY", "invoice_company": "Company Name", '
        '"invoice_amount": "Amount", "invoice_currency": "Currency", "comments": "Any issues or comments"}\n\n'
        f"Invoice Text:\n{text}"
    )

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
        max_tokens=300,
    )

    raw_output = response.choices[0].message.content

    try:
        return json.loads(raw_output)
    except json.JSONDecodeError:
        import re

        match = re.search(r"{.*}", raw_output, re.DOTALL)
        if match:
            return json.loads(match.group())
        raise ValueError(f"Invalid JSON in response: {raw_output}")


def extract_text_from_pdf(pdf_path):
    text = ""
    with fitz.open(pdf_path) as doc:
        for page in doc:
            text += page.get_text()
    return text


def process_pdfs_to_excel(pdf_folder, excel_file="Invoice_Record.xlsx"):
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Date", "Company", "Amount", "Currency", "Comments"])

    for filename in os.listdir(pdf_folder):
        if filename.lower().endswith(".pdf"):
            file_path = os.path.join(pdf_folder, filename)
            try:
                text = extract_text_from_pdf(file_path)
                data = generate_invoice_summary(text)

                sheet.append(
                    [
                        data["invoice_date"],
                        data["invoice_company"],
                        data["invoice_amount"],
                        data["invoice_currency"],
                        data.get("comments", ""),
                    ]
                )

                print(f"Processed: {filename}")
            except Exception as e:
                print(f"Failed to process {filename}: {e}")

    wb.save(excel_file)
    print(f"\n✅ Invoice data saved to: {excel_file}")


# === Example Usage ===
if __name__ == "__main__":
    folder_with_pdfs = "Invoices/"
    process_pdfs_to_excel(folder_with_pdfs)
